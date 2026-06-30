#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 工具箱 — 公共工具模块
提供文件检测、安全读写、格式化输出等基础能力。
"""

import os
import sys
import json
import csv
import shutil
import importlib
from pathlib import Path
from datetime import datetime
from io import StringIO


# ============================================================
# 依赖检测
# ============================================================

REQUIRED_PACKAGES = {
    "openpyxl": "openpyxl>=3.1.0",
    "pandas": "pandas>=2.0.0",
    "xlsxwriter": "xlsxwriter>=3.1.0",
    "xlrd": "xlrd>=2.0.0",
    "msoffcrypto": "msoffcrypto-tool>=5.0.0",
    "numpy": "numpy>=1.24.0",
}


def ensure_dependencies(auto_install=True):
    """检测必要依赖是否已安装，缺失时自动安装。

    Args:
        auto_install: 是否自动安装缺失依赖（默认 True）

    Returns:
        list: 最终仍缺失的包名列表，为空表示全部就绪
    """
    missing = []
    for pkg_import, pkg_spec in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(pkg_import)
        except ImportError:
            missing.append(pkg_spec)

    if not missing:
        return []

    if auto_install:
        print(f"[自动安装] 检测到缺失依赖，正在安装: {', '.join(missing)}", file=sys.stderr)
        import subprocess
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "--quiet", "--disable-pip-version-check"] + missing,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
                timeout=120
            )
            print("[自动安装] ✅ 依赖安装完成", file=sys.stderr)
            # 验证安装结果
            still_missing = []
            for pkg_import in REQUIRED_PACKAGES:
                try:
                    importlib.import_module(pkg_import)
                except ImportError:
                    still_missing.append(pkg_import)
            if still_missing:
                print(f"[警告] 以下依赖安装后仍无法导入: {', '.join(still_missing)}", file=sys.stderr)
            return still_missing
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
            print(f"[警告] 自动安装失败: {e}", file=sys.stderr)
            print(f"[提示] 请手动执行: pip install {' '.join(missing)}", file=sys.stderr)
            return [pkg.split('>')[0].split('=')[0] for pkg in missing]
    else:
        print(f"[警告] 缺少以下依赖: {', '.join(missing)}", file=sys.stderr)
        print(f"[提示] 请执行: pip install {' '.join(missing)}", file=sys.stderr)
        return [pkg.split('>')[0].split('=')[0] for pkg in missing]


# ============================================================
# 文件类型检测
# ============================================================

def detect_file_type(path):
    """根据文件扩展名和内容特征检测文件类型。

    Args:
        path: 文件路径

    Returns:
        str: 文件类型字符串 ('xlsx', 'xls', 'csv', 'json', 'tsv', 'unknown')

    Raises:
        FileNotFoundError: 文件不存在时抛出
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    # 先根据扩展名判断
    ext = path.suffix.lower()
    ext_map = {
        ".xlsx": "xlsx",
        ".xls": "xls",
        ".csv": "csv",
        ".json": "json",
        ".tsv": "tsv",
    }
    if ext in ext_map:
        return ext_map[ext]

    # 尝试读取文件头部特征
    try:
        with open(path, "rb") as f:
            header = f.read(8)
        # xlsx 是 ZIP 格式
        if header[:4] == b"PK\x03\x04":
            return "xlsx"
        # xls 是 OLE2 格式
        if header[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            return "xls"
    except Exception:
        pass

    # 尝试作为文本文件判断
    try:
        with open(path, "r", encoding="utf-8") as f:
            first_line = f.readline(4096)
        # JSON 通常以 [ 或 { 开头
        stripped = first_line.strip()
        if stripped and stripped[0] in ("{", "["):
            return "json"
        # 含制表符较多可能是 TSV
        if "\t" in first_line and first_line.count("\t") > first_line.count(","):
            return "tsv"
        # 默认视为 CSV
        if "," in first_line:
            return "csv"
    except Exception:
        pass

    return "unknown"


# ============================================================
# 大文件检测
# ============================================================

def is_large_file(path, threshold=100000):
    """判断文件是否为大文件（基于行数估算）。

    对于 Excel 文件使用 openpyxl read_only 模式快速获取行数；
    对于文本文件通过抽样估算总行数避免全量扫描。

    Args:
        path: 文件路径
        threshold: 行数阈值，默认 10 万行

    Returns:
        bool: 超过阈值返回 True
    """
    path = Path(path)
    file_type = detect_file_type(path)

    if file_type == "xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                if ws.max_row and ws.max_row > threshold:
                    wb.close()
                    return True
            wb.close()
            return False
        except Exception:
            pass

    # 对文本类文件用文件大小估算（假设平均每行 100 字节）
    file_size = path.stat().st_size
    estimated_lines = file_size / 100
    if estimated_lines > threshold * 2:
        return True

    # 如果文件不太大，精确计数
    if file_size < 50 * 1024 * 1024:  # 50MB 以下精确计数
        try:
            with open(path, "rb") as f:
                line_count = sum(1 for _ in f)
            return line_count > threshold
        except Exception:
            pass

    return estimated_lines > threshold


# ============================================================
# 安全读取
# ============================================================

def safe_read(path, **kwargs):
    """安全读取文件，大文件自动使用流式/只读模式。

    支持 xlsx/xls/csv/json 格式，异常行跳过并记录到 ErrorCollector。

    Args:
        path: 文件路径
        **kwargs: 传递给 pandas.read_excel / read_csv 的额外参数
            - sheet_name: Sheet 名称或索引
            - nrows: 读取行数
            - usecols: 使用的列
            - header: 表头行号

    Returns:
        tuple: (DataFrame 或 dict of DataFrames, ErrorCollector)
    """
    import pandas as pd

    path = Path(path)
    file_type = detect_file_type(path)
    errors = ErrorCollector()
    large = is_large_file(path)

    # sheet_name=None 会让 pandas 返回 dict，默认用第一个 Sheet
    if 'sheet_name' in kwargs and kwargs['sheet_name'] is None:
        kwargs['sheet_name'] = 0

    if large:
        errors.add_warning(f"检测到大文件，自动启用流式读取模式")

    try:
        if file_type in ("xlsx", "xls"):
            # 大文件使用 openpyxl 的 read_only 模式
            engine = "openpyxl" if file_type == "xlsx" else "xlrd"
            read_kwargs = {"engine": engine}
            read_kwargs.update(kwargs)

            if large and "nrows" not in read_kwargs:
                # 大文件默认只读前 10 万行，避免内存溢出
                read_kwargs["nrows"] = 100000
                errors.add_warning("大文件截断：仅读取前 100000 行")

            try:
                df = pd.read_excel(path, **read_kwargs)
            except Exception as e:
                # read_excel 不支持 on_bad_lines，直接报错
                errors.add_error(f"读取 Excel 失败: {e}")
                df = pd.DataFrame()

        elif file_type in ("csv", "tsv"):
            sep = "\t" if file_type == "tsv" else ","
            read_kwargs = {"sep": sep, "on_bad_lines": "skip"}
            read_kwargs.update(kwargs)

            if large and "nrows" not in read_kwargs:
                read_kwargs["nrows"] = 100000
                errors.add_warning("大文件截断：仅读取前 100000 行")

            # 尝试多种编码
            for encoding in ("utf-8", "gbk", "gb2312", "latin-1"):
                try:
                    df = pd.read_csv(path, encoding=encoding, **read_kwargs)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                df = pd.read_csv(path, encoding="latin-1", **read_kwargs)
                errors.add_warning("编码检测失败，使用 latin-1 兜底")

        elif file_type == "json":
            try:
                df = pd.read_json(path, **kwargs)
            except ValueError:
                # 可能是 JSON Lines 格式
                df = pd.read_json(path, lines=True, **kwargs)

        else:
            raise ValueError(f"不支持的文件类型: {file_type}")

        return df, errors

    except Exception as e:
        errors.add_error(f"文件读取失败: {e}")
        raise


# ============================================================
# 安全写入
# ============================================================

def safe_write(wb, path):
    """安全写入 Excel 文件，写入前自动备份原文件。

    Args:
        wb: openpyxl Workbook 对象
        path: 输出文件路径

    Returns:
        str: 写入的文件路径
    """
    path = Path(path)

    # 如果目标文件已存在，先备份
    if path.exists():
        backup_name = f"{path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        backup_path = path.parent / backup_name
        try:
            shutil.copy2(path, backup_path)
            print(f"[信息] 已备份原文件到: {backup_path}", file=sys.stderr)
        except Exception as e:
            print(f"[警告] 备份失败: {e}", file=sys.stderr)

    # 确保输出目录存在
    path.parent.mkdir(parents=True, exist_ok=True)

    # 写入文件
    try:
        wb.save(str(path))
        return str(path)
    except Exception as e:
        raise IOError(f"文件写入失败: {e}")


# ============================================================
# 输出格式化
# ============================================================

def format_output(data, fmt="table"):
    """统一输出格式化，支持多种输出格式。

    Args:
        data: pandas DataFrame 或可转换为 DataFrame 的数据
        fmt: 输出格式 ('json', 'csv', 'markdown', 'table', 'tsv')

    Returns:
        str: 格式化后的字符串
    """
    import pandas as pd

    if not isinstance(data, pd.DataFrame):
        data = pd.DataFrame(data)

    if fmt == "json":
        return data.to_json(orient="records", force_ascii=False, indent=2)

    elif fmt == "csv":
        return data.to_csv(index=False)

    elif fmt == "tsv":
        return data.to_csv(index=False, sep="\t")

    elif fmt == "markdown":
        try:
            return data.to_markdown(index=False)
        except ImportError:
            return data.to_string(index=False)

    elif fmt == "table":
        # 使用 tabulate 风格的表格输出
        try:
            return data.to_markdown(index=False)
        except ImportError:
            return data.to_string(index=False)

    else:
        raise ValueError(f"不支持的输出格式: {fmt}，可选: json/csv/markdown/table/tsv")


# ============================================================
# 错误收集器
# ============================================================

class ErrorCollector:
    """收集处理过程中的错误和警告，最后汇总报告。

    用法:
        errors = ErrorCollector()
        errors.add_error("行10: 数据格式错误")
        errors.add_warning("编码自动转换为 UTF-8")
        if errors.has_errors():
            print(errors.report())
    """

    def __init__(self):
        self._errors = []
        self._warnings = []

    def add_error(self, msg):
        """添加一条错误记录。"""
        self._errors.append(msg)

    def add_warning(self, msg):
        """添加一条警告记录。"""
        self._warnings.append(msg)

    def has_errors(self):
        """是否有错误。"""
        return len(self._errors) > 0

    def raise_if_errors(self):
        """如果有错误则抛出异常。"""
        if self.has_errors():
            raise RuntimeError('\n'.join(self._errors))

    def has_warnings(self):
        """是否有警告。"""
        return len(self._warnings) > 0

    def has_issues(self):
        """是否有错误或警告。"""
        return self.has_errors() or self.has_warnings()

    @property
    def error_count(self):
        return len(self._errors)

    @property
    def warning_count(self):
        return len(self._warnings)

    def report(self):
        """生成汇总报告字符串。"""
        lines = []
        if self._errors:
            lines.append(f"❌ 错误 ({len(self._errors)} 条):")
            for i, e in enumerate(self._errors, 1):
                lines.append(f"  {i}. {e}")
        if self._warnings:
            lines.append(f"⚠️  警告 ({len(self._warnings)} 条):")
            for i, w in enumerate(self._warnings, 1):
                lines.append(f"  {i}. {w}")
        if not lines:
            lines.append("✅ 处理完成，无错误或警告。")
        return "\n".join(lines)

    def __str__(self):
        return self.report()

    def __repr__(self):
        return f"ErrorCollector(errors={self.error_count}, warnings={self.warning_count})"
