#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
差异对比工具 - 对比两个 Excel 文件的差异

场景：对比两版花名册，找出新增/离职/信息变更
输出：新增行、删除行、修改行（具体字段变化）+ 统计摘要
可选：生成带颜色标注的高亮 Excel
"""

import argparse
import json
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import detect_file_type, safe_read, format_output, ErrorCollector

# 高亮颜色定义
FILL_ADDED = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 绿色 = 新增
FILL_DELETED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # 红色 = 删除
FILL_MODIFIED = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 黄色 = 修改
FONT_HEADER = Font(bold=True)


def compare_dataframes(old_df: pd.DataFrame, new_df: pd.DataFrame,
                       key_col: str, ignore_columns: list = None) -> dict:
    """对比两个 DataFrame，返回差异详情"""
    # 过滤忽略列
    compare_cols = [c for c in old_df.columns if c in new_df.columns]
    if ignore_columns:
        compare_cols = [c for c in compare_cols if c not in ignore_columns]

    # 确保主键列在对比列中
    if key_col not in compare_cols:
        compare_cols.insert(0, key_col)

    # 获取主键集合
    old_keys = set(old_df[key_col].dropna().astype(str))
    new_keys = set(new_df[key_col].dropna().astype(str))

    added_keys = new_keys - old_keys
    deleted_keys = old_keys - new_keys
    common_keys = old_keys & new_keys

    # 新增行
    added_rows = []
    for key in sorted(added_keys):
        row = new_df[new_df[key_col].astype(str) == key].iloc[0]
        added_rows.append({col: str(row.get(col, '')) for col in compare_cols})

    # 删除行
    deleted_rows = []
    for key in sorted(deleted_keys):
        row = old_df[old_df[key_col].astype(str) == key].iloc[0]
        deleted_rows.append({col: str(row.get(col, '')) for col in compare_cols})

    # 修改行
    modified_rows = []
    for key in sorted(common_keys):
        old_row = old_df[old_df[key_col].astype(str) == key].iloc[0]
        new_row = new_df[new_df[key_col].astype(str) == key].iloc[0]

        changes = []
        for col in compare_cols:
            if col == key_col:
                continue
            old_val = str(old_row.get(col, '')) if pd.notna(old_row.get(col)) else ''
            new_val = str(new_row.get(col, '')) if pd.notna(new_row.get(col)) else ''
            if old_val != new_val:
                changes.append({
                    'column': col,
                    'old_value': old_val,
                    'new_value': new_val,
                })

        if changes:
            modified_rows.append({
                'key': key,
                'changes': changes,
            })

    return {
        'added': added_rows,
        'deleted': deleted_rows,
        'modified': modified_rows,
        'summary': {
            'old_rows': len(old_df),
            'new_rows': len(new_df),
            'added_count': len(added_rows),
            'deleted_count': len(deleted_rows),
            'modified_count': len(modified_rows),
            'unchanged_count': len(common_keys) - len(modified_rows),
        }
    }


def generate_highlight_excel(diff_result: dict, old_df: pd.DataFrame,
                             new_df: pd.DataFrame, key_col: str,
                             output_path: str):
    """生成带颜色标注的 Excel 差异报告"""
    wb = Workbook()

    # Sheet 1: 摘要
    ws_summary = wb.active
    ws_summary.title = '差异摘要'
    summary = diff_result['summary']
    summary_data = [
        ['差异对比报告', ''],
        ['', ''],
        ['旧文件行数', summary['old_rows']],
        ['新文件行数', summary['new_rows']],
        ['新增行数', summary['added_count']],
        ['删除行数', summary['deleted_count']],
        ['修改行数', summary['modified_count']],
        ['未变化行数', summary['unchanged_count']],
    ]
    for row_data in summary_data:
        ws_summary.append(row_data)
    ws_summary['A1'].font = FONT_HEADER

    # Sheet 2: 新增行（绿色）
    if diff_result['added']:
        ws_added = wb.create_sheet('新增')
        headers = list(diff_result['added'][0].keys())
        ws_added.append(headers)
        for cell in ws_added[1]:
            cell.font = FONT_HEADER
        for row in diff_result['added']:
            row_data = [row.get(h, '') for h in headers]
            ws_added.append(row_data)
            for cell in ws_added[ws_added.max_row]:
                cell.fill = FILL_ADDED

    # Sheet 3: 删除行（红色）
    if diff_result['deleted']:
        ws_deleted = wb.create_sheet('删除')
        headers = list(diff_result['deleted'][0].keys())
        ws_deleted.append(headers)
        for cell in ws_deleted[1]:
            cell.font = FONT_HEADER
        for row in diff_result['deleted']:
            row_data = [row.get(h, '') for h in headers]
            ws_deleted.append(row_data)
            for cell in ws_deleted[ws_deleted.max_row]:
                cell.fill = FILL_DELETED

    # Sheet 4: 修改行（黄色）
    if diff_result['modified']:
        ws_modified = wb.create_sheet('修改')
        ws_modified.append(['主键', '字段', '旧值', '新值'])
        for cell in ws_modified[1]:
            cell.font = FONT_HEADER
        for item in diff_result['modified']:
            for change in item['changes']:
                ws_modified.append([
                    item['key'], change['column'],
                    change['old_value'], change['new_value']
                ])
                for cell in ws_modified[ws_modified.max_row]:
                    cell.fill = FILL_MODIFIED

    wb.save(output_path)


def format_text_report(diff_result: dict) -> str:
    """格式化为文本报告"""
    lines = []
    summary = diff_result['summary']

    lines.append("=" * 60)
    lines.append("📊 差异对比报告")
    lines.append("=" * 60)
    lines.append(f"旧文件: {summary['old_rows']} 行")
    lines.append(f"新文件: {summary['new_rows']} 行")
    lines.append(f"")
    lines.append(f"🟢 新增: {summary['added_count']} 行")
    lines.append(f"🔴 删除: {summary['deleted_count']} 行")
    lines.append(f"🟡 修改: {summary['modified_count']} 行")
    lines.append(f"⚪ 未变: {summary['unchanged_count']} 行")

    # 新增详情
    if diff_result['added']:
        lines.append(f"\n{'─' * 60}")
        lines.append("🟢 新增行:")
        for i, row in enumerate(diff_result['added'][:20], 1):
            row_str = ' | '.join([f"{k}={v}" for k, v in list(row.items())[:5]])
            lines.append(f"  {i}. {row_str}")
        if len(diff_result['added']) > 20:
            lines.append(f"  ... 还有 {len(diff_result['added']) - 20} 行")

    # 删除详情
    if diff_result['deleted']:
        lines.append(f"\n{'─' * 60}")
        lines.append("🔴 删除行:")
        for i, row in enumerate(diff_result['deleted'][:20], 1):
            row_str = ' | '.join([f"{k}={v}" for k, v in list(row.items())[:5]])
            lines.append(f"  {i}. {row_str}")
        if len(diff_result['deleted']) > 20:
            lines.append(f"  ... 还有 {len(diff_result['deleted']) - 20} 行")

    # 修改详情
    if diff_result['modified']:
        lines.append(f"\n{'─' * 60}")
        lines.append("🟡 修改行:")
        for i, item in enumerate(diff_result['modified'][:20], 1):
            changes_str = ', '.join([
                f"{c['column']}: {c['old_value']} → {c['new_value']}"
                for c in item['changes'][:3]
            ])
            lines.append(f"  {i}. [{item['key']}] {changes_str}")
            if len(item['changes']) > 3:
                lines.append(f"     ... 还有 {len(item['changes']) - 3} 个字段变化")
        if len(diff_result['modified']) > 20:
            lines.append(f"  ... 还有 {len(diff_result['modified']) - 20} 行")

    lines.append(f"\n{'=' * 60}")
    return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(
        description='Excel 差异对比工具 - 对比两个文件的差异'
    )
    parser.add_argument('--old', required=True, help='旧文件路径')
    parser.add_argument('--new', required=True, help='新文件路径')
    parser.add_argument('--key', required=True, help='主键列名（用于行匹配）')
    parser.add_argument('--output', help='差异报告输出路径')
    parser.add_argument('--format', choices=['text', 'json'], default='text',
                        help='报告格式（默认 text）')
    parser.add_argument('--highlight', help='生成高亮标注 Excel 的输出路径')
    parser.add_argument('--ignore-columns', help='忽略的列（逗号分隔）')
    parser.add_argument('--sheet', help='指定 Sheet 名称')

    args = parser.parse_args()

    # 校验输入
    if not os.path.exists(args.old):
        print(f"❌ 旧文件不存在: {args.old}")
        sys.exit(1)
    if not os.path.exists(args.new):
        print(f"❌ 新文件不存在: {args.new}")
        sys.exit(1)

    # 加载数据
    try:
        old_df, _err1 = safe_read(args.old, sheet_name=args.sheet)
        new_df, _err2 = safe_read(args.new, sheet_name=args.sheet)
    except Exception as e:
        print(f"❌ 读取文件失败: {e}")
        sys.exit(1)

    # 验证主键列
    if args.key not in old_df.columns:
        print(f"❌ 主键列 '{args.key}' 在旧文件中不存在")
        sys.exit(1)
    if args.key not in new_df.columns:
        print(f"❌ 主键列 '{args.key}' 在新文件中不存在")
        sys.exit(1)

    # 解析忽略列
    ignore_cols = []
    if args.ignore_columns:
        ignore_cols = [c.strip() for c in args.ignore_columns.split(',')]

    print(f"📊 旧文件: {args.old} ({len(old_df)} 行)")
    print(f"📊 新文件: {args.new} ({len(new_df)} 行)")
    print(f"🔑 主键列: {args.key}")
    if ignore_cols:
        print(f"🚫 忽略列: {', '.join(ignore_cols)}")
    print()

    # 执行对比
    diff_result = compare_dataframes(old_df, new_df, args.key, ignore_cols)

    # 输出文本/JSON 报告
    if args.format == 'json':
        report_text = json.dumps(diff_result, ensure_ascii=False, indent=2)
    else:
        report_text = format_text_report(diff_result)

    if args.output:
        os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else '.', exist_ok=True)
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(report_text)
        print(f"✅ 差异报告已保存: {args.output}")
    else:
        print(report_text)

    # 生成高亮 Excel
    if args.highlight:
        os.makedirs(os.path.dirname(args.highlight) if os.path.dirname(args.highlight) else '.', exist_ok=True)
        generate_highlight_excel(diff_result, old_df, new_df, args.key, args.highlight)
        print(f"✅ 高亮 Excel 已保存: {args.highlight}")


if __name__ == "__main__":
    main()
