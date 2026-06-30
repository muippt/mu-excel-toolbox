#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 工具箱 — 格式互转
支持 xlsx↔csv↔json↔markdown↔tsv 双向转换。
"""

import argparse
import sys
import json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from utils import (
    detect_file_type,
    ensure_dependencies,
    safe_read,
    is_large_file,
    ErrorCollector,
)


# 支持的目标格式
SUPPORTED_FORMATS = ("xlsx", "csv", "json", "markdown", "tsv", "md")

# 扩展名到格式映射
EXT_TO_FORMAT = {
    ".xlsx": "xlsx",
    ".xls": "xlsx",  # 输出统一用 xlsx
    ".csv": "csv",
    ".json": "json",
    ".md": "markdown",
    ".markdown": "markdown",
    ".tsv": "tsv",
    ".txt": "tsv",
}


def infer_output_format(output_path):
    """从输出文件路径推断目标格式。

    Args:
        output_path: 输出文件路径

    Returns:
        str or None: 推断的格式，无法推断时返回 None
    """
    ext = Path(output_path).suffix.lower()
    return EXT_TO_FORMAT.get(ext)


def generate_output_path(input_path, target_format):
    """根据输入路径和目标格式生成默认输出路径。

    Args:
        input_path: 输入文件路径
        target_format: 目标格式

    Returns:
        Path: 输出文件路径
    """
    input_path = Path(input_path)
    format_ext = {
        "xlsx": ".xlsx",
        "csv": ".csv",
        "json": ".json",
        "markdown": ".md",
        "md": ".md",
        "tsv": ".tsv",
    }
    ext = format_ext.get(target_format, f".{target_format}")
    return input_path.with_suffix(ext)


def convert_to_xlsx(df, output_path, sheet_name="Sheet1"):
    """将 DataFrame 转换为 xlsx 文件。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写入表头
    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 写入数据
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            # 处理 NaN 值
            import pandas as pd
            if pd.isna(value):
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动列宽
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, min(ws.max_row + 1, 51)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell_str = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in cell_str)
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(str(output_path))
    return str(output_path)


def convert_to_csv(df, output_path, encoding="utf-8", sep=","):
    """将 DataFrame 转换为 CSV 文件。"""
    df.to_csv(output_path, index=False, encoding=encoding, sep=sep)
    return str(output_path)


def convert_to_json(df, output_path, encoding="utf-8"):
    """将 DataFrame 转换为 JSON 文件。"""
    # 处理 NaN 值
    df_clean = df.where(df.notna(), None)
    records = df_clean.to_dict(orient="records")

    with open(output_path, "w", encoding=encoding) as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)
    return str(output_path)


def convert_to_markdown(df, output_path, encoding="utf-8"):
    """将 DataFrame 转换为 Markdown 表格文件。"""
    try:
        md_content = df.to_markdown(index=False)
    except ImportError:
        # tabulate 未安装时用 to_string 兜底
        md_content = df.to_string(index=False)
    with open(output_path, "w", encoding=encoding) as f:
        f.write(md_content)
    return str(output_path)


def convert_to_tsv(df, output_path, encoding="utf-8"):
    """将 DataFrame 转换为 TSV 文件。"""
    df.to_csv(output_path, index=False, encoding=encoding, sep="\t")
    return str(output_path)


def convert_file(input_path, target_format, output_path=None,
                 sheet_name=None, encoding="utf-8", sep=None):
    """执行文件格式转换。

    Args:
        input_path: 输入文件路径
        target_format: 目标格式
        output_path: 输出路径（可选，自动生成）
        sheet_name: Sheet 名称（读取时指定或写入时使用）
        encoding: 输出编码
        sep: CSV 分隔符（仅对 CSV 输出有效）

    Returns:
        tuple: (输出路径, ErrorCollector)
    """
    import pandas as pd

    errors = ErrorCollector()
    input_path = Path(input_path)

    # 确定输出路径
    if output_path is None:
        output_path = generate_output_path(input_path, target_format)
    else:
        output_path = Path(output_path)

    # 确保输出目录存在
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 标准化格式名
    if target_format == "md":
        target_format = "markdown"

    # 读取源文件
    read_kwargs = {}
    if sheet_name is not None:
        try:
            read_kwargs["sheet_name"] = int(sheet_name)
        except (ValueError, TypeError):
            read_kwargs["sheet_name"] = sheet_name

    df, read_errors = safe_read(input_path, **read_kwargs)
    errors._errors.extend(read_errors._errors)
    errors._warnings.extend(read_errors._warnings)

    # 转换输出
    if target_format == "xlsx":
        s_name = sheet_name if sheet_name else "Sheet1"
        result = convert_to_xlsx(df, output_path, sheet_name=s_name)
    elif target_format == "csv":
        csv_sep = sep if sep else ","
        result = convert_to_csv(df, output_path, encoding=encoding, sep=csv_sep)
    elif target_format == "json":
        result = convert_to_json(df, output_path, encoding=encoding)
    elif target_format == "markdown":
        result = convert_to_markdown(df, output_path, encoding=encoding)
    elif target_format == "tsv":
        result = convert_to_tsv(df, output_path, encoding=encoding)
    else:
        raise ValueError(f"不支持的目标格式: {target_format}，可选: {', '.join(SUPPORTED_FORMATS)}")

    return result, errors


def main():
    parser = argparse.ArgumentParser(
        description="Excel/CSV/JSON/Markdown 格式互转",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # Excel 转 CSV
  python convert.py data.xlsx --to csv

  # CSV 转 Excel
  python convert.py data.csv --to xlsx

  # Excel 转 JSON
  python convert.py data.xlsx --to json --output result.json

  # JSON 转 Markdown
  python convert.py data.json --to markdown

  # 指定 Sheet 转换
  python convert.py data.xlsx --to csv --sheet "销售数据"

  # 自动推断格式（从 --output 扩展名）
  python convert.py data.xlsx --output result.csv

  # 指定编码和分隔符
  python convert.py data.csv --to xlsx --encoding gbk --sep ";"
        """,
    )
    parser.add_argument("input_file", help="输入文件路径")
    parser.add_argument("--to", "-t", choices=["xlsx", "csv", "json", "markdown", "md", "tsv"],
                        help="目标格式")
    parser.add_argument("--output", "-o", help="输出文件路径（可选，自动生成）")
    parser.add_argument("--sheet", "-s", help="Sheet 名称或索引")
    parser.add_argument("--encoding", "-e", default="utf-8", help="输出编码（默认: utf-8）")
    parser.add_argument("--sep", help="CSV 分隔符（默认: 逗号）")

    args = parser.parse_args()

    # 检查依赖
    missing = ensure_dependencies()
    if missing:
        sys.exit(1)

    # 检查文件
    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"❌ 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    # 确定目标格式
    target_format = args.to
    if target_format is None and args.output:
        target_format = infer_output_format(args.output)
    if target_format is None:
        print("❌ 请指定目标格式（--to）或提供带扩展名的输出路径（--output）", file=sys.stderr)
        sys.exit(1)

    # 检查是否同格式转换
    source_type = detect_file_type(input_path)
    if source_type == target_format:
        print(f"⚠️  源文件已经是 {target_format} 格式", file=sys.stderr)

    try:
        result_path, errors = convert_file(
            input_path,
            target_format,
            output_path=args.output,
            sheet_name=args.sheet,
            encoding=args.encoding,
            sep=args.sep,
        )

        print(f"✅ 转换完成: {input_path} → {result_path}")
        print(f"   格式: {source_type} → {target_format}")

        if errors.has_issues():
            print(f"\n{errors.report()}", file=sys.stderr)

    except Exception as e:
        print(f"❌ 转换失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
