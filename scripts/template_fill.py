#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模板填充工具 - 用模板 + 数据批量生成 Excel 文件

场景：用员工数据批量生成 offer letter、合同、通知书等
占位符语法：{{字段名}}，支持嵌套 {{部门.名称}}
"""

import argparse
import json
import os
import re
import sys
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import detect_file_type, safe_read, format_output, ErrorCollector

# 占位符正则：匹配 {{字段名}} 或 {{父级.子级}}
PLACEHOLDER_PATTERN = re.compile(r'\{\{([^}]+)\}\}')


def resolve_nested_value(data: dict, key_path: str):
    """解析嵌套字段值，如 '部门.名称' -> data['部门']['名称']"""
    keys = key_path.strip().split('.')
    value = data
    for k in keys:
        if isinstance(value, dict) and k in value:
            value = value[k]
        else:
            return None
    return value


def fill_cell_value(cell_value, record: dict) -> str:
    """替换单元格中的所有占位符"""
    if not isinstance(cell_value, str):
        return cell_value

    def replacer(match):
        key_path = match.group(1)
        val = resolve_nested_value(record, key_path)
        if val is None:
            return match.group(0)  # 未找到则保留原占位符
        return str(val)

    return PLACEHOLDER_PATTERN.sub(replacer, cell_value)


def generate_filename(naming_template: str, record: dict, index: int) -> str:
    """根据命名规则生成输出文件名"""
    if not naming_template:
        return f"output_{index + 1}.xlsx"

    def replacer(match):
        key_path = match.group(1)
        val = resolve_nested_value(record, key_path)
        if val is None:
            return f"未知_{key_path}"
        # 移除文件名中的非法字符
        val_str = str(val)
        val_str = re.sub(r'[\\/:*?"<>|]', '_', val_str)
        return val_str

    filename = PLACEHOLDER_PATTERN.sub(replacer, naming_template)
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    return filename


def load_data(data_path: str) -> list:
    """加载数据文件，返回记录列表（每条记录为 dict）"""
    file_type = detect_file_type(data_path)

    if file_type == 'json':
        with open(data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        elif isinstance(data, dict):
            return [data]
        else:
            raise ValueError("JSON 数据格式不支持，需要数组或对象")
    elif file_type in ('csv', 'xlsx', 'xls'):
        df, _errors = safe_read(data_path)
        # 将 NaN 转为 None，再转为字典列表
        df = df.where(pd.notnull(df), None)
        return df.to_dict('records')
    else:
        raise ValueError(f"不支持的数据文件格式: {data_path}")


def fill_template(template_path: str, record: dict) -> 'Workbook':
    """用一条记录填充模板，返回填充后的 workbook"""
    wb = load_workbook(template_path)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                    cell.value = fill_cell_value(cell.value, record)

    return wb


def batch_fill(template_path: str, records: list, output_dir: str, naming: str):
    """批量模式：每条记录生成一个文件"""
    os.makedirs(output_dir, exist_ok=True)
    generated = []

    for i, record in enumerate(records):
        wb = fill_template(template_path, record)
        filename = generate_filename(naming, record, i)
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)
        generated.append(output_path)
        print(f"  ✅ [{i + 1}/{len(records)}] {filename}")

    return generated


def single_fill(template_path: str, records: list, output_dir: str, naming: str):
    """单文件模式：所有数据填入同一文件的不同 Sheet"""
    os.makedirs(output_dir, exist_ok=True)

    # 加载模板作为基础
    base_wb = load_workbook(template_path)
    template_ws = base_wb.active

    # 删除所有现有 sheet，重新创建
    for sheet_name in base_wb.sheetnames:
        del base_wb[sheet_name]

    for i, record in enumerate(records):
        # 每条记录创建一个新 sheet
        sheet_name = generate_filename(naming, record, i).replace('.xlsx', '')[:31]  # Sheet名最长31字符
        ws = base_wb.create_sheet(title=sheet_name)

        # 从模板复制并填充
        temp_wb = load_workbook(template_path)
        temp_ws = temp_wb.active

        for row in temp_ws.iter_rows():
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column)
                if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                    new_cell.value = fill_cell_value(cell.value, record)
                else:
                    new_cell.value = cell.value
                # 复制样式
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy(cell.alignment)

        # 复制列宽
        for col_idx, col_dim in temp_ws.column_dimensions.items():
            ws.column_dimensions[col_idx].width = col_dim.width

        print(f"  ✅ [{i + 1}/{len(records)}] Sheet: {sheet_name}")

    output_filename = naming if naming and naming.endswith('.xlsx') else "output_all.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    base_wb.save(output_path)
    print(f"\n📄 所有数据已合并到: {output_path}")
    return [output_path]


def main():
    parser = argparse.ArgumentParser(
        description='Excel 模板填充工具 - 用模板 + 数据批量生成 Excel 文件'
    )
    parser.add_argument('--template', required=True, help='模板文件路径（含 {{占位符}}）')
    parser.add_argument('--data', required=True, help='数据文件路径（JSON/CSV/Excel）')
    parser.add_argument('--output-dir', default='./output', help='输出目录（默认 ./output）')
    parser.add_argument('--naming', default='', help='输出文件命名规则，如 {{姓名}}_offer.xlsx')
    parser.add_argument('--single', action='store_true', help='单文件模式：所有数据填入同一文件的不同 Sheet')

    args = parser.parse_args()

    # 校验输入
    if not os.path.exists(args.template):
        print(f"❌ 模板文件不存在: {args.template}")
        sys.exit(1)
    if not os.path.exists(args.data):
        print(f"❌ 数据文件不存在: {args.data}")
        sys.exit(1)

    print(f"📋 模板: {args.template}")
    print(f"📊 数据: {args.data}")

    # 加载数据
    try:
        records = load_data(args.data)
    except Exception as e:
        print(f"❌ 加载数据失败: {e}")
        sys.exit(1)

    print(f"📝 共 {len(records)} 条记录\n")

    if len(records) == 0:
        print("⚠️ 数据为空，无需处理")
        sys.exit(0)

    # 执行填充
    if args.single:
        print("📦 单文件模式（合并到不同 Sheet）")
        generated = single_fill(args.template, records, args.output_dir, args.naming)
    else:
        print("📦 批量模式（每条数据一个文件）")
        generated = batch_fill(args.template, records, args.output_dir, args.naming)

    print(f"\n✅ 完成！共生成 {len(generated)} 个文件，输出目录: {args.output_dir}")


if __name__ == "__main__":
    main()
