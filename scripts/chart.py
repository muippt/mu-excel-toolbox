#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 图表生成工具
功能：在 Excel 中生成各类图表（柱状图、折线图、饼图、散点图、组合图等）
使用 xlsxwriter 渲染图表（效果优于 openpyxl），先用 openpyxl 读取已有文件数据
"""

import argparse
import os
import sys
from typing import Optional, List, Dict

# 添加脚本所在目录到 path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import detect_file_type, safe_read, format_output, ErrorCollector

try:
    import xlsxwriter
except ImportError:
    print("❌ 缺少 xlsxwriter 库，请运行: pip install xlsxwriter")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("❌ 缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("❌ 缺少 pandas 库，请运行: pip install pandas")
    sys.exit(1)


# 图表类型映射
CHART_TYPE_MAP = {
    'bar': 'column',       # xlsxwriter 中 column = 柱状图
    'column': 'column',
    'line': 'line',
    'pie': 'pie',
    'scatter': 'scatter',
    'area': 'area',
    'doughnut': 'doughnut',
    'radar': 'radar',
    'combo': 'combo',
}


def col_to_index(col: str) -> int:
    """列标识转 0-based 索引（xlsxwriter 使用 0-based）"""
    try:
        return int(col) - 1  # 用户输入 1-based 转为 0-based
    except ValueError:
        return column_index_from_string(col.upper()) - 1


def parse_position(pos_str: str) -> tuple:
    """解析图表位置字符串如 'E2' 为 (row, col) 0-based"""
    import re
    match = re.match(r'^([A-Za-z]+)(\d+)$', pos_str)
    if match:
        col = column_index_from_string(match.group(1).upper()) - 1
        row = int(match.group(2)) - 1
        return (row, col)
    return (1, 4)  # 默认 E2


def parse_size(size_str: str) -> tuple:
    """解析尺寸字符串 'width,height'"""
    parts = size_str.split(',')
    if len(parts) == 2:
        return (int(parts[0].strip()), int(parts[1].strip()))
    return (480, 300)  # 默认尺寸


def read_data(input_file: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """读取 Excel/CSV 数据"""
    df, _errors = safe_read(input_file, sheet_name=sheet_name)
    return df


def create_chart(args, collector: ErrorCollector):
    """创建图表并写入新的 Excel 文件"""
    # 读取源数据
    df = read_data(args.input_file)

    if df.empty:
        collector.add_error("输入文件无数据")
        collector.raise_if_errors()

    # 解析列参数
    x_col_name = args.x_column
    y_col_names = [c.strip() for c in args.y_columns.split(',')]

    # 验证列名存在
    all_columns = list(df.columns)
    if x_col_name not in all_columns:
        # 尝试作为列索引（字母）
        try:
            x_idx = col_to_index(x_col_name)
            x_col_name = all_columns[x_idx]
        except (ValueError, IndexError):
            collector.add_error(f"X轴列 '{args.x_column}' 不存在。可用列: {all_columns}")
            collector.raise_if_errors()

    resolved_y_cols = []
    for y_col in y_col_names:
        if y_col in all_columns:
            resolved_y_cols.append(y_col)
        else:
            try:
                y_idx = col_to_index(y_col)
                resolved_y_cols.append(all_columns[y_idx])
            except (ValueError, IndexError):
                collector.add_error(f"Y轴列 '{y_col}' 不存在。可用列: {all_columns}")

    if collector.has_errors():
        collector.raise_if_errors()

    y_col_names = resolved_y_cols

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        base, _ext = os.path.splitext(args.input_file)
        output_path = f"{base}_chart.xlsx"
    if not output_path.endswith('.xlsx'):
        output_path += '.xlsx'

    # 创建 xlsxwriter 工作簿
    workbook = xlsxwriter.Workbook(output_path)

    # 写入数据到 data sheet
    data_sheet = workbook.add_worksheet('Data')

    # 写表头
    headers = [x_col_name] + y_col_names
    # 获取相关列的数据
    write_cols = [x_col_name] + y_col_names
    sub_df = df[write_cols].copy()

    for col_idx, header in enumerate(headers):
        data_sheet.write(0, col_idx, header)

    # 写数据
    for row_idx in range(len(sub_df)):
        for col_idx, col_name in enumerate(write_cols):
            value = sub_df.iloc[row_idx][col_name]
            # 处理 NaN
            if pd.isna(value):
                data_sheet.write_blank(row_idx + 1, col_idx, None)
            elif isinstance(value, (int, float)):
                data_sheet.write_number(row_idx + 1, col_idx, value)
            else:
                data_sheet.write_string(row_idx + 1, col_idx, str(value))

    num_rows = len(sub_df)

    # 创建图表 sheet
    chart_sheet_name = args.sheet or 'Chart'
    chart_ws = workbook.add_worksheet(chart_sheet_name)

    # 确定图表类型
    chart_type = args.type.lower()

    if chart_type == 'combo' and args.combo:
        # 组合图表
        chart = _create_combo_chart(workbook, args, y_col_names, num_rows, collector)
    else:
        xlsw_type = CHART_TYPE_MAP.get(chart_type, 'column')
        chart = workbook.add_chart({'type': xlsw_type})

        # 添加数据系列
        for i, y_col in enumerate(y_col_names):
            col_idx = i + 1  # Y 列从第 2 列开始（0-based = 1）
            series_config = {
                'name': ['Data', 0, col_idx],
                'categories': ['Data', 1, 0, num_rows, 0],
                'values': ['Data', 1, col_idx, num_rows, col_idx],
            }

            # 饼图只用第一个系列
            if xlsw_type == 'pie' and i > 0:
                collector.add_warning(f"饼图只支持一个数据系列，已忽略: {y_col}")
                break

            chart.add_series(series_config)

    # 设置图表属性
    if args.title:
        chart.set_title({'name': args.title})

    # 设置图例位置
    if args.legend:
        legend_pos_map = {
            'top': 'top',
            'bottom': 'bottom',
            'left': 'left',
            'right': 'right',
            'none': 'none',
        }
        pos = legend_pos_map.get(args.legend.lower(), 'right')
        if pos == 'none':
            chart.set_legend({'none': True})
        else:
            chart.set_legend({'position': pos})

    # 设置尺寸
    if args.size:
        width, height = parse_size(args.size)
        chart.set_size({'width': width, 'height': height})

    # 插入图表
    if args.position:
        pos_row, pos_col = parse_position(args.position)
        chart_ws.insert_chart(pos_row, pos_col, chart)
    else:
        chart_ws.insert_chart(1, 0, chart)

    workbook.close()
    return output_path


def _create_combo_chart(workbook, args, y_col_names, num_rows, collector):
    """创建组合图表（柱状+折线）"""
    # 解析 combo 参数，格式: "bar:col1,col2;line:col3"
    combo_config = {}
    parts = args.combo.split(';')
    for part in parts:
        if ':' in part:
            chart_type, cols = part.split(':', 1)
            combo_config[chart_type.strip()] = [c.strip() for c in cols.split(',')]
        else:
            collector.add_warning(f"组合图表配置格式错误: {part}，应为 type:col1,col2")

    # 主图表用第一种类型
    primary_type = list(combo_config.keys())[0] if combo_config else 'column'
    xlsw_primary = CHART_TYPE_MAP.get(primary_type, 'column')
    chart = workbook.add_chart({'type': xlsw_primary})

    # 添加主类型系列
    primary_cols = combo_config.get(primary_type, [])
    for col_name in primary_cols:
        if col_name in y_col_names:
            col_idx = y_col_names.index(col_name) + 1
            chart.add_series({
                'name': ['Data', 0, col_idx],
                'categories': ['Data', 1, 0, num_rows, 0],
                'values': ['Data', 1, col_idx, num_rows, col_idx],
            })

    # 添加其他类型系列
    for chart_type, cols in combo_config.items():
        if chart_type == primary_type:
            continue
        xlsw_type = CHART_TYPE_MAP.get(chart_type, 'line')
        for col_name in cols:
            if col_name in y_col_names:
                col_idx = y_col_names.index(col_name) + 1
                chart.add_series({
                    'name': ['Data', 0, col_idx],
                    'categories': ['Data', 1, 0, num_rows, 0],
                    'values': ['Data', 1, col_idx, num_rows, col_idx],
                    'y2_axis': True,
                })
                # 创建叠加图
                secondary_chart = workbook.add_chart({'type': xlsw_type})
                secondary_chart.add_series({
                    'name': ['Data', 0, col_idx],
                    'categories': ['Data', 1, 0, num_rows, 0],
                    'values': ['Data', 1, col_idx, num_rows, col_idx],
                })
                chart.combine(secondary_chart)

    return chart


def main():
    parser = argparse.ArgumentParser(
        description='Excel 图表生成工具（基于 xlsxwriter）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 柱状图
  python chart.py sales.xlsx --type bar --x-column 月份 --y-columns 销售额 --title "月度销售"

  # 折线图（多系列）
  python chart.py data.xlsx --type line --x-column 日期 --y-columns "收入,支出" --legend bottom

  # 饼图
  python chart.py data.xlsx --type pie --x-column 部门 --y-columns 占比 --title "部门占比"

  # 组合图（柱状+折线）
  python chart.py data.xlsx --type combo --x-column 月份 --y-columns "销售额,增长率" \\
      --combo "bar:销售额;line:增长率" --title "销售趋势"

  # 指定位置和大小
  python chart.py data.xlsx --type bar --x-column A --y-columns B,C \\
      --position E2 --size 600,400 --sheet Report
"""
    )

    parser.add_argument('input_file', help='输入 Excel/CSV 文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径（默认在原文件名后加 _chart）')
    parser.add_argument('--type', '-t', default='bar',
                        choices=['bar', 'column', 'line', 'pie', 'scatter', 'area',
                                 'combo', 'doughnut', 'radar'],
                        help='图表类型')
    parser.add_argument('--title', help='图表标题')
    parser.add_argument('--x-column', required=True, help='X轴数据列（列名或列字母）')
    parser.add_argument('--y-columns', required=True, help='Y轴数据列（逗号分隔，支持多列）')
    parser.add_argument('--sheet', help='图表放置的 Sheet 名称（默认 Chart）')
    parser.add_argument('--position', help='图表位置（如 E2，默认 A2）')
    parser.add_argument('--size', help='图表尺寸（width,height 像素，如 600,400）')
    parser.add_argument('--legend', choices=['top', 'bottom', 'left', 'right', 'none'],
                        help='图例位置')
    parser.add_argument('--combo', help='组合图表配置（格式: bar:col1,col2;line:col3）')

    args = parser.parse_args()

    # 验证输入
    if not os.path.exists(args.input_file):
        print(f"❌ 文件不存在: {args.input_file}")
        sys.exit(1)

    if args.type == 'combo' and not args.combo:
        print("❌ 组合图表必须指定 --combo 参数（格式: bar:col1;line:col2）")
        sys.exit(1)

    collector = ErrorCollector()

    try:
        output_path = create_chart(args, collector)
        print(f"✅ 图表已生成: {output_path}")
    except RuntimeError as e:
        print(f"❌ {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 图表生成失败: {e}")
        sys.exit(1)

    if collector.has_warnings():
        print(collector.report())


if __name__ == "__main__":
    main()
