#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 函数公式批量填充工具
功能：批量填充 Excel 公式，支持自定义公式模板、内置快捷公式、公式验证
"""

import argparse
import os
import sys
import re
from typing import Optional, List

# 添加脚本所在目录到 path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import detect_file_type, safe_read, format_output, ErrorCollector

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("❌ 缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)


def col_to_index(col: str) -> int:
    """列标识转 1-based 索引，支持字母或数字"""
    try:
        return int(col)
    except ValueError:
        return column_index_from_string(col.upper())


def col_to_letter(col: str) -> str:
    """列标识统一转为字母"""
    try:
        return get_column_letter(int(col))
    except ValueError:
        return col.upper()


def parse_range(range_str: str, max_row: int) -> tuple:
    """
    解析范围字符串，返回 (start_row, end_row)
    支持格式: "2:100", "2:", ":50", "all"
    """
    if not range_str or range_str.lower() == 'all':
        return (2, max_row)

    parts = range_str.split(':')
    if len(parts) == 2:
        start = int(parts[0]) if parts[0] else 2
        end = int(parts[1]) if parts[1] else max_row
        return (start, end)
    else:
        # 单个数字表示从该行到末尾
        return (int(parts[0]), max_row)


def validate_formula_refs(formula: str, max_row: int, max_col: int, collector: ErrorCollector):
    """验证公式中的单元格引用是否超出数据范围"""
    # 匹配单元格引用如 A1, $B$2, C100
    ref_pattern = r'\$?([A-Z]+)\$?(\d+)'
    refs = re.findall(ref_pattern, formula.upper())

    for col_ref, row_ref in refs:
        row_num = int(row_ref)
        try:
            col_num = column_index_from_string(col_ref)
        except Exception:
            continue

        if row_num > max_row:
            collector.add_warning(
                f"公式引用行 {row_num} 超出数据范围(最大行: {max_row})，公式: {formula}"
            )
        if col_num > max_col:
            collector.add_warning(
                f"公式引用列 {col_ref}({col_num}) 超出数据范围(最大列: {max_col})，公式: {formula}"
            )


def apply_formula(wb, args, collector: ErrorCollector):
    """应用自定义公式模板"""
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    target_col_idx = col_to_index(args.column)
    start_row, end_row = parse_range(args.range, max_row)

    # 确保范围合理
    end_row = min(end_row, max_row)

    for row in range(start_row, end_row + 1):
        # 替换 {row} 占位符为实际行号
        formula = args.formula.replace('{row}', str(row))

        # 验证公式引用
        validate_formula_refs(formula, max_row, max_col, collector)

        # 确保公式以 = 开头
        if not formula.startswith('='):
            formula = '=' + formula

        ws.cell(row=row, column=target_col_idx, value=formula)

    # 自动求和
    if args.auto_sum:
        sum_row = end_row + 1
        col_letter = get_column_letter(target_col_idx)
        ws.cell(row=sum_row, column=target_col_idx,
                value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")

    # 自动平均
    if args.auto_average:
        avg_row = end_row + (2 if args.auto_sum else 1)
        col_letter = get_column_letter(target_col_idx)
        ws.cell(row=avg_row, column=target_col_idx,
                value=f"=AVERAGE({col_letter}{start_row}:{col_letter}{end_row})")

    return wb


def apply_sum_columns(wb, args, collector: ErrorCollector):
    """快捷：多列求和填入目标列"""
    ws = wb.active
    max_row = ws.max_row

    target_col_idx = col_to_index(args.column)
    source_cols = [col_to_letter(c.strip()) for c in args.sum_columns.split(',')]
    start_row, end_row = parse_range(args.range, max_row)
    end_row = min(end_row, max_row)

    for row in range(start_row, end_row + 1):
        refs = '+'.join([f"{c}{row}" for c in source_cols])
        ws.cell(row=row, column=target_col_idx, value=f"={refs}")

    # 末尾汇总
    if args.auto_sum:
        col_letter = get_column_letter(target_col_idx)
        ws.cell(row=end_row + 1, column=target_col_idx,
                value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")

    return wb


def apply_average_columns(wb, args, collector: ErrorCollector):
    """快捷：多列平均填入目标列"""
    ws = wb.active
    max_row = ws.max_row

    target_col_idx = col_to_index(args.column)
    source_cols = [col_to_letter(c.strip()) for c in args.average_columns.split(',')]
    start_row, end_row = parse_range(args.range, max_row)
    end_row = min(end_row, max_row)

    for row in range(start_row, end_row + 1):
        refs = ','.join([f"{c}{row}" for c in source_cols])
        ws.cell(row=row, column=target_col_idx, value=f"=AVERAGE({refs})")

    return wb


def apply_count_if(wb, args, collector: ErrorCollector):
    """快捷：条件计数"""
    ws = wb.active
    max_row = ws.max_row

    target_col_idx = col_to_index(args.column)
    # count_if 格式: "源列,条件" 如 "C,>60"
    parts = args.count_if.split(',', 1)
    if len(parts) != 2:
        collector.add_error("--count-if 格式应为: 源列,条件 (如 C,>60)")
        return wb

    source_col = col_to_letter(parts[0].strip())
    criteria = parts[1].strip()

    # 在目标列第一个数据行写入 COUNTIF
    start_row, end_row = parse_range(args.range, max_row)
    formula = f'=COUNTIF({source_col}{start_row}:{source_col}{end_row},"{criteria}")'
    ws.cell(row=start_row, column=target_col_idx, value=formula)

    return wb


def apply_vlookup(wb, args, collector: ErrorCollector):
    """快捷：生成 VLOOKUP 公式"""
    ws = wb.active
    max_row = ws.max_row

    target_col_idx = col_to_index(args.column)
    # vlookup_formula 格式: "查找列,表范围,返回列号,精确匹配"
    # 如 "A,Sheet2!A:D,3,0"
    parts = args.vlookup_formula.split(',')
    if len(parts) < 3:
        collector.add_error("--vlookup-formula 格式: 查找列,表范围,返回列号[,匹配类型]")
        return wb

    lookup_col = col_to_letter(parts[0].strip())
    table_range = parts[1].strip()
    col_index = parts[2].strip()
    match_type = parts[3].strip() if len(parts) > 3 else '0'

    start_row, end_row = parse_range(args.range, max_row)
    end_row = min(end_row, max_row)

    for row in range(start_row, end_row + 1):
        formula = f"=VLOOKUP({lookup_col}{row},{table_range},{col_index},{match_type})"
        ws.cell(row=row, column=target_col_idx, value=formula)

    return wb


def main():
    parser = argparse.ArgumentParser(
        description='Excel 函数公式批量填充工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 自定义公式：在 E 列填入 B+C+D 的公式
  python formula.py data.xlsx --column E --formula "=B{row}+C{row}+D{row}"

  # 快捷求和：B~D 列求和填入 E 列
  python formula.py data.xlsx --column E --sum-columns B,C,D

  # 快捷平均：B~D 列平均填入 F 列
  python formula.py data.xlsx --column F --average-columns B,C,D

  # VLOOKUP：在 D 列生成 VLOOKUP 公式
  python formula.py data.xlsx --column D --vlookup-formula "A,Sheet2!A:C,3,0"

  # 自动在末尾加 SUM
  python formula.py data.xlsx --column E --formula "=B{row}+C{row}" --auto-sum

  # 条件计数
  python formula.py data.xlsx --column F --count-if "D,>60"
"""
    )

    parser.add_argument('input_file', help='输入 Excel 文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径（默认覆盖原文件）')
    parser.add_argument('--sheet', '-s', help='Sheet 名称（默认第一个）')
    parser.add_argument('--column', '-c', required=True, help='目标列（字母或数字）')
    parser.add_argument('--formula', '-f', help='公式模板（用 {row} 占位行号）')
    parser.add_argument('--range', '-r', help='应用范围（如 2:100，默认全部数据行）')
    parser.add_argument('--auto-sum', action='store_true', help='自动在末尾加 SUM 汇总')
    parser.add_argument('--auto-average', action='store_true', help='自动在末尾加 AVERAGE')

    # 快捷公式
    shortcut_group = parser.add_argument_group('快捷公式')
    shortcut_group.add_argument('--sum-columns', help='指定列求和（逗号分隔，如 B,C,D）')
    shortcut_group.add_argument('--average-columns', help='指定列求平均（逗号分隔）')
    shortcut_group.add_argument('--count-if', help='条件计数（格式: 源列,条件 如 C,>60）')
    shortcut_group.add_argument('--vlookup-formula',
                                help='VLOOKUP 公式（格式: 查找列,表范围,返回列号[,匹配类型]）')

    args = parser.parse_args()

    # 验证输入
    if not os.path.exists(args.input_file):
        print(f"❌ 文件不存在: {args.input_file}")
        sys.exit(1)

    # 至少需要一种公式指定方式
    if not any([args.formula, args.sum_columns, args.average_columns,
                args.count_if, args.vlookup_formula]):
        print("❌ 请指定公式：--formula / --sum-columns / --average-columns / --count-if / --vlookup-formula")
        sys.exit(1)

    collector = ErrorCollector()

    # 加载工作簿
    try:
        wb = load_workbook(args.input_file)
        if hasattr(args, 'sheet') and args.sheet:
            if args.sheet in wb.sheetnames:
                wb.active = wb[args.sheet]
            else:
                print(f"❌ Sheet '{args.sheet}' 不存在，可用: {wb.sheetnames}")
                sys.exit(1)
    except Exception as e:
        print(f"❌ 无法打开文件: {e}")
        sys.exit(1)

    # 根据指定的公式类型执行
    if args.formula:
        wb = apply_formula(wb, args, collector)
    elif args.sum_columns:
        wb = apply_sum_columns(wb, args, collector)
    elif args.average_columns:
        wb = apply_average_columns(wb, args, collector)
    elif args.count_if:
        wb = apply_count_if(wb, args, collector)
    elif args.vlookup_formula:
        wb = apply_vlookup(wb, args, collector)

    # 保存
    output_path = args.output or args.input_file
    try:
        wb.save(output_path)
        print(f"✅ 已保存到: {output_path}")
    except Exception as e:
        print(f"❌ 保存失败: {e}")
        sys.exit(1)

    # 输出错误/警告报告
    if collector.has_errors() or collector.has_warnings():
        print(collector.report())


if __name__ == "__main__":
    main()
