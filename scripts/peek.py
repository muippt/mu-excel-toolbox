#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 工具箱 — 智能预览
快速了解 Excel/CSV 文件的结构：Sheet 列表、列信息、数据类型推断、前 N 行预览。
"""

import argparse
import sys
import json
from pathlib import Path

# 添加脚本目录到路径
sys.path.insert(0, str(Path(__file__).parent))
from utils import detect_file_type, ensure_dependencies, format_output, ErrorCollector


def peek_xlsx(path, nrows=5):
    """预览 xlsx 文件结构。

    Args:
        path: 文件路径
        nrows: 预览行数

    Returns:
        dict: 包含各 Sheet 信息的字典
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    result = {"file": str(path), "sheets": []}

    for ws in wb.worksheets:
        sheet_info = {
            "name": ws.title,
            "rows": ws.max_row or 0,
            "columns": ws.max_column or 0,
            "headers": [],
            "dtypes": [],
            "preview": [],
        }

        # 读取表头和前 N 行数据
        rows_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > nrows:  # 表头 + nrows 行数据
                break
            rows_data.append(list(row))

        if rows_data:
            # 第一行作为表头
            sheet_info["headers"] = [str(h) if h is not None else f"列{j+1}" for j, h in enumerate(rows_data[0])]

            # 数据行
            data_rows = rows_data[1:]
            sheet_info["preview"] = data_rows

            # 推断数据类型
            if data_rows:
                for col_idx in range(len(sheet_info["headers"])):
                    col_values = [row[col_idx] for row in data_rows if col_idx < len(row) and row[col_idx] is not None]
                    dtype = _infer_type(col_values)
                    sheet_info["dtypes"].append(dtype)
            else:
                sheet_info["dtypes"] = ["unknown"] * len(sheet_info["headers"])

        result["sheets"].append(sheet_info)

    wb.close()
    return result


def peek_xls(path, nrows=5):
    """预览 xls 文件结构。"""
    import xlrd

    wb = xlrd.open_workbook(path)
    result = {"file": str(path), "sheets": []}

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        sheet_info = {
            "name": ws.name,
            "rows": ws.nrows,
            "columns": ws.ncols,
            "headers": [],
            "dtypes": [],
            "preview": [],
        }

        if ws.nrows > 0:
            # 第一行作为表头
            sheet_info["headers"] = [str(ws.cell_value(0, c)) or f"列{c+1}" for c in range(ws.ncols)]

            # 预览数据行
            end_row = min(ws.nrows, nrows + 1)
            for r in range(1, end_row):
                row_data = [ws.cell_value(r, c) for c in range(ws.ncols)]
                sheet_info["preview"].append(row_data)

            # 推断数据类型
            if sheet_info["preview"]:
                for col_idx in range(ws.ncols):
                    col_values = [row[col_idx] for row in sheet_info["preview"] if col_idx < len(row) and row[col_idx] not in (None, "")]
                    sheet_info["dtypes"].append(_infer_type(col_values))
            else:
                sheet_info["dtypes"] = ["unknown"] * ws.ncols

        result["sheets"].append(sheet_info)

    return result


def peek_csv(path, nrows=5):
    """预览 CSV/TSV 文件结构。"""
    import pandas as pd

    # 检测分隔符和编码
    file_type = detect_file_type(path)
    sep = "\t" if file_type == "tsv" else ","

    # 尝试多种编码
    df = None
    for encoding in ("utf-8", "gbk", "gb2312", "latin-1"):
        try:
            df = pd.read_csv(path, sep=sep, nrows=nrows, encoding=encoding, on_bad_lines="skip")
            break
        except UnicodeDecodeError:
            continue

    if df is None:
        raise ValueError(f"无法读取文件: {path}")

    # 估算总行数（不全量加载）
    total_rows = _estimate_rows(path)

    sheet_info = {
        "name": "Sheet1",
        "rows": total_rows,
        "columns": len(df.columns),
        "headers": list(df.columns.astype(str)),
        "dtypes": [str(dt) for dt in df.dtypes],
        "preview": df.values.tolist(),
    }

    return {"file": str(path), "sheets": [sheet_info]}


def _estimate_rows(path):
    """估算文本文件总行数（不全量加载大文件）。"""
    path = Path(path)
    file_size = path.stat().st_size

    # 小文件精确计数
    if file_size < 10 * 1024 * 1024:  # 10MB
        with open(path, "rb") as f:
            return sum(1 for _ in f) - 1  # 减去表头行

    # 大文件通过采样估算
    sample_size = 1024 * 100  # 100KB
    with open(path, "rb") as f:
        sample = f.read(sample_size)
    lines_in_sample = sample.count(b"\n")
    if lines_in_sample == 0:
        return 1
    avg_line_size = sample_size / lines_in_sample
    estimated = int(file_size / avg_line_size) - 1
    return estimated


def _infer_type(values):
    """推断一组值的数据类型。"""
    if not values:
        return "unknown"

    type_counts = {"int": 0, "float": 0, "date": 0, "str": 0, "bool": 0}

    for v in values:
        if isinstance(v, bool):
            type_counts["bool"] += 1
        elif isinstance(v, int):
            type_counts["int"] += 1
        elif isinstance(v, float):
            type_counts["float"] += 1
        elif isinstance(v, str):
            # 尝试判断是否为数字字符串
            try:
                int(v)
                type_counts["int"] += 1
                continue
            except (ValueError, TypeError):
                pass
            try:
                float(v)
                type_counts["float"] += 1
                continue
            except (ValueError, TypeError):
                pass
            type_counts["str"] += 1
        else:
            type_counts["str"] += 1

    # 返回占比最高的类型
    dominant = max(type_counts, key=type_counts.get)
    if type_counts[dominant] == 0:
        return "unknown"
    return dominant


def format_peek_result(result, fmt="table"):
    """格式化预览结果输出。"""
    import pandas as pd

    output_lines = []
    output_lines.append(f"📁 文件: {result['file']}")
    output_lines.append(f"📊 Sheet 数量: {len(result['sheets'])}")
    output_lines.append("")

    for sheet in result["sheets"]:
        output_lines.append(f"━━━ Sheet: {sheet['name']} ━━━")
        output_lines.append(f"  行数: {sheet['rows']} | 列数: {sheet['columns']}")
        output_lines.append("")

        # 列信息表
        if sheet["headers"]:
            col_info = []
            for i, header in enumerate(sheet["headers"]):
                dtype = sheet["dtypes"][i] if i < len(sheet["dtypes"]) else "unknown"
                col_info.append({"#": i + 1, "列名": header, "类型": dtype})

            output_lines.append("  列信息:")
            col_df = pd.DataFrame(col_info)
            for line in col_df.to_string(index=False).split("\n"):
                output_lines.append(f"    {line}")
            output_lines.append("")

        # 数据预览
        if sheet["preview"]:
            output_lines.append("  数据预览:")
            preview_df = pd.DataFrame(sheet["preview"], columns=sheet["headers"])
            if fmt == "json":
                output_lines.append(preview_df.to_json(orient="records", force_ascii=False, indent=2))
            else:
                for line in preview_df.to_string(index=False).split("\n"):
                    output_lines.append(f"    {line}")
        output_lines.append("")

    return "\n".join(output_lines)


def main():
    parser = argparse.ArgumentParser(
        description="智能预览 Excel/CSV 文件结构",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python peek.py data.xlsx
  python peek.py data.csv --rows 10
  python peek.py report.xlsx --format json
        """,
    )
    parser.add_argument("input_file", help="要预览的文件路径")
    parser.add_argument("--rows", "-r", type=int, default=5, help="预览行数（默认: 5）")
    parser.add_argument("--format", "-f", choices=["table", "json", "markdown"], default="table", help="输出格式（默认: table）")

    args = parser.parse_args()

    # 检查依赖
    missing = ensure_dependencies()
    if missing:
        sys.exit(1)

    # 检查文件是否存在
    path = Path(args.input_file)
    if not path.exists():
        print(f"❌ 文件不存在: {path}", file=sys.stderr)
        sys.exit(1)

    # 检测文件类型并预览
    try:
        file_type = detect_file_type(path)

        if file_type == "xlsx":
            result = peek_xlsx(path, nrows=args.rows)
        elif file_type == "xls":
            result = peek_xls(path, nrows=args.rows)
        elif file_type in ("csv", "tsv"):
            result = peek_csv(path, nrows=args.rows)
        else:
            print(f"❌ 不支持的文件类型: {file_type}", file=sys.stderr)
            sys.exit(1)

        # 输出结果
        if args.format == "json":
            print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        else:
            print(format_peek_result(result, fmt=args.format))

    except Exception as e:
        print(f"❌ 预览失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
