#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
图片插入工具 - 在 Excel 单元格中插入图片

场景：在每个员工对应行插入照片、产品图片等
支持批量插入（按目录下文件名匹配）
自动调整行高适应图片
"""

import argparse
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter, column_index_from_string

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import detect_file_type, safe_read, format_output, ErrorCollector

# 支持的图片格式
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.tif'}


def parse_size(size_str: str) -> tuple:
    """解析尺寸参数，格式：width,height（单位像素）"""
    if not size_str:
        return (100, 100)
    parts = size_str.split(',')
    if len(parts) == 2:
        return (int(parts[0].strip()), int(parts[1].strip()))
    elif len(parts) == 1:
        val = int(parts[0].strip())
        return (val, val)
    else:
        return (100, 100)


def parse_column(col_str: str) -> int:
    """解析列参数，支持字母（如 D）或数字（如 4）"""
    try:
        return int(col_str)
    except ValueError:
        return column_index_from_string(col_str.upper())


def find_images_in_dir(image_dir: str) -> dict:
    """扫描目录下所有图片文件，返回 {文件名(无后缀): 完整路径}"""
    image_map = {}
    if not os.path.isdir(image_dir):
        return image_map

    for f in os.listdir(image_dir):
        ext = os.path.splitext(f)[1].lower()
        if ext in IMAGE_EXTENSIONS:
            name_without_ext = os.path.splitext(f)[0]
            image_map[name_without_ext] = os.path.join(image_dir, f)

    return image_map


def pixels_to_row_height(pixels: int) -> float:
    """将像素高度转换为 Excel 行高（磅）"""
    # Excel 行高单位是磅（point），1磅 ≈ 1.33像素
    return pixels * 0.75


def pixels_to_col_width(pixels: int) -> float:
    """将像素宽度转换为 Excel 列宽（字符宽度）"""
    # Excel 列宽单位约为 7像素/字符
    return pixels / 7.0


def insert_images(input_file: str, output_file: str, column: int,
                  images: list, match_by: str, match_col_idx: int,
                  size: tuple, start_row: int, image_dir: str = None):
    """在 Excel 中插入图片"""
    wb = load_workbook(input_file)
    ws = wb.active

    width, height = size
    col_letter = get_column_letter(column)
    inserted_count = 0
    errors = []

    # 设置目标列宽（适应图片宽度 + 边距）
    ws.column_dimensions[col_letter].width = pixels_to_col_width(width + 10)

    # 构建图片匹配映射
    if image_dir:
        # 按目录匹配模式
        image_map = find_images_in_dir(image_dir)
    elif images:
        # 按列表模式
        image_map = {}
        for img_path in images:
            if os.path.isfile(img_path):
                name = os.path.splitext(os.path.basename(img_path))[0]
                image_map[name] = img_path
    else:
        print("❌ 未指定图片来源（--images 或目录）")
        return 0, ["未指定图片来源"]

    if not image_map:
        print("⚠️ 未找到任何图片文件")
        return 0, ["未找到任何图片文件"]

    print(f"📁 找到 {len(image_map)} 张图片")
    print(f"📐 图片尺寸: {width}×{height} 像素")
    print(f"📍 插入位置: 第 {col_letter} 列，从第 {start_row} 行开始\n")

    # 逐行处理
    for row_num in range(start_row, ws.max_row + 1):
        # 获取匹配值
        if match_col_idx:
            match_cell = ws.cell(row=row_num, column=match_col_idx)
            match_value = str(match_cell.value).strip() if match_cell.value else ''
        else:
            # 无匹配列时，按行序依次插入
            idx = row_num - start_row
            if idx < len(images):
                img_path = images[idx]
                if os.path.isfile(img_path):
                    try:
                        img = XlImage(img_path)
                        img.width = width
                        img.height = height
                        anchor = f"{col_letter}{row_num}"
                        ws.add_image(img, anchor)
                        ws.row_dimensions[row_num].height = pixels_to_row_height(height + 5)
                        inserted_count += 1
                        print(f"  ✅ 行 {row_num}: {os.path.basename(img_path)}")
                    except Exception as e:
                        errors.append(f"行 {row_num}: {e}")
            continue

        if not match_value:
            continue

        # 在 image_map 中查找匹配的图片
        img_path = image_map.get(match_value)
        if not img_path:
            # 尝试模糊匹配（去掉空格等）
            clean_match = match_value.replace(' ', '').replace('_', '')
            for name, path in image_map.items():
                if name.replace(' ', '').replace('_', '') == clean_match:
                    img_path = path
                    break

        if img_path and os.path.isfile(img_path):
            try:
                img = XlImage(img_path)
                img.width = width
                img.height = height
                anchor = f"{col_letter}{row_num}"
                ws.add_image(img, anchor)

                # 调整行高适应图片
                ws.row_dimensions[row_num].height = pixels_to_row_height(height + 5)
                inserted_count += 1
                print(f"  ✅ 行 {row_num}: {match_value} → {os.path.basename(img_path)}")
            except Exception as e:
                errors.append(f"行 {row_num} ({match_value}): {e}")
                print(f"  ❌ 行 {row_num}: {match_value} → 插入失败: {e}")
        else:
            # 未匹配到图片（不算错误，可能该行就没图片）
            pass

    # 保存
    wb.save(output_file)
    return inserted_count, errors


def main():
    parser = argparse.ArgumentParser(
        description='Excel 图片插入工具 - 在单元格中批量插入图片'
    )
    parser.add_argument('input_file', help='输入 Excel 文件路径')
    parser.add_argument('--output', help='输出文件路径（不指定则覆盖原文件）')
    parser.add_argument('--column', default='D', help='图片放置列（字母或数字，默认 D）')
    parser.add_argument('--images', nargs='*', help='图片路径列表或目录路径')
    parser.add_argument('--match-by', help='按哪列的值匹配图片文件名（列名）')
    parser.add_argument('--size', default='100,100', help='图片尺寸 width,height（像素，默认 100,100）')
    parser.add_argument('--start-row', type=int, default=2, help='起始行号（默认 2，跳过表头）')
    parser.add_argument('--sheet', help='指定 Sheet 名称')

    args = parser.parse_args()

    # 校验输入
    if not os.path.exists(args.input_file):
        print(f"❌ 输入文件不存在: {args.input_file}")
        sys.exit(1)

    output_file = args.output or args.input_file
    target_col = parse_column(args.column)
    size = parse_size(args.size)

    # 确定图片来源
    image_dir = None
    image_list = []
    if args.images:
        if len(args.images) == 1 and os.path.isdir(args.images[0]):
            image_dir = args.images[0]
        else:
            image_list = args.images

    # 确定匹配列
    match_col_idx = None
    if args.match_by:
        # 读取文件获取列名 → 列索引
        wb_temp = load_workbook(args.input_file, read_only=True)
        ws_temp = wb_temp.active
        header_row = [cell.value for cell in ws_temp[1]]
        wb_temp.close()

        if args.match_by in header_row:
            match_col_idx = header_row.index(args.match_by) + 1
        else:
            print(f"❌ 匹配列 '{args.match_by}' 不存在于文件中")
            print(f"   可用列: {', '.join([str(h) for h in header_row if h])}")
            sys.exit(1)

    print(f"📄 输入: {args.input_file}")
    print(f"📄 输出: {output_file}")

    # 执行插入
    inserted, errors = insert_images(
        input_file=args.input_file,
        output_file=output_file,
        column=target_col,
        images=image_list,
        match_by=args.match_by,
        match_col_idx=match_col_idx,
        size=size,
        start_row=args.start_row,
        image_dir=image_dir,
    )

    print(f"\n✅ 完成！成功插入 {inserted} 张图片")
    if errors:
        print(f"⚠️ {len(errors)} 个错误:")
        for err in errors[:10]:
            print(f"  - {err}")
    print(f"📄 输出文件: {output_file}")


if __name__ == "__main__":
    main()
