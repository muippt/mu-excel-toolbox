#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
密码保护/解密工具 - 给 Excel 加密或解密

场景1：给薪酬表加密码保护（文件级 + Sheet级）
场景2：解密加密的 Excel 文件
注意：
  - openpyxl 提供 Sheet 级保护（防止编辑单元格）
  - 文件级加密（打开时需要密码）需要 msoffcrypto-tool
"""

import argparse
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import detect_file_type, safe_read, format_output, ErrorCollector


def encrypt_file(input_file: str, output_file: str, password: str):
    """文件级加密（使用 msoffcrypto-tool）"""
    try:
        import msoffcrypto
    except ImportError:
        print("❌ 缺少依赖: msoffcrypto-tool")
        print("   请运行: pip install msoffcrypto-tool")
        sys.exit(1)

    with open(input_file, 'rb') as f_in:
        file = msoffcrypto.OfficeFile(f_in)
        # 如果原文件本身就是加密的，先解密
        if file.is_encrypted():
            print("⚠️ 原文件已加密，将先解密再重新加密")
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
            try:
                file.load_key(password='')  # 尝试空密码
                file.decrypt(open(temp_path, 'wb'))
            except Exception:
                os.unlink(temp_path)
                print("❌ 原文件已加密且无法用空密码解密，请先解密后再加密")
                sys.exit(1)
            input_file = temp_path

    # 加密
    with open(input_file, 'rb') as f_in:
        file = msoffcrypto.OfficeFile(f_in)
        file.load_key(password=password)
        with open(output_file, 'wb') as f_out:
            file.encrypt(password, f_out)

    # 清理临时文件
    if 'temp_path' in locals() and os.path.exists(temp_path):
        os.unlink(temp_path)

    print(f"✅ 文件级加密完成: {output_file}")
    print(f"🔑 密码: {'*' * len(password)} (长度{len(password)})")
    print(f"⚠️  请妥善保管密码，忘记密码将无法恢复文件")


def decrypt_file(input_file: str, output_file: str, password: str):
    """文件级解密（使用 msoffcrypto-tool）"""
    try:
        import msoffcrypto
    except ImportError:
        print("❌ 缺少依赖: msoffcrypto-tool")
        print("   请运行: pip install msoffcrypto-tool")
        sys.exit(1)

    with open(input_file, 'rb') as f_in:
        file = msoffcrypto.OfficeFile(f_in)
        if not file.is_encrypted():
            print("ℹ️ 文件未加密（文件级），直接复制")
            import shutil
            shutil.copy2(input_file, output_file)
            return

        file.load_key(password=password)
        try:
            with open(output_file, 'wb') as f_out:
                file.decrypt(f_out)
            print(f"✅ 文件级解密完成: {output_file}")
        except Exception as e:
            print(f"❌ 解密失败（密码可能不正确）: {e}")
            sys.exit(1)


def protect_sheets(input_file: str, output_file: str, password: str,
                   sheet_names: list = None, allow_operations: list = None):
    """Sheet 级保护（防止编辑单元格）"""
    from openpyxl import load_workbook

    wb = load_workbook(input_file)

    # 确定要保护的 Sheet
    if sheet_names:
        target_sheets = [ws for ws in wb.worksheets if ws.title in sheet_names]
        if not target_sheets:
            print(f"⚠️ 未找到指定的 Sheet: {sheet_names}")
            print(f"   可用 Sheet: {[ws.title for ws in wb.worksheets]}")
            sys.exit(1)
    else:
        target_sheets = wb.worksheets  # 保护所有 Sheet

    # 设置允许的操作
    protection_kwargs = {}
    if allow_operations:
        for op in allow_operations:
            if op == 'select_cells':
                protection_kwargs['select_locked_cells'] = False
                protection_kwargs['select_unlocked_cells'] = False
            elif op == 'sort':
                protection_kwargs['sort'] = False
            elif op == 'filter':
                protection_kwargs['autoFilter'] = False
            elif op == 'insert_rows':
                protection_kwargs['insertRows'] = False
            elif op == 'delete_rows':
                protection_kwargs['deleteRows'] = False
            elif op == 'format_cells':
                protection_kwargs['formatCells'] = False

    for ws in target_sheets:
        ws.protection.sheet = True
        ws.protection.password = password
        # 设置允许操作
        for key, val in protection_kwargs.items():
            setattr(ws.protection, key, val)
        print(f"  🔒 Sheet '{ws.title}' 已保护")

    wb.save(output_file)
    print(f"\n✅ Sheet 级保护完成: {output_file}")
    print(f"🔑 密码: {'*' * len(password)} (长度{len(password)})")
    if allow_operations:
        print(f"📋 允许操作: {', '.join(allow_operations)}")


def unprotect_sheets(input_file: str, output_file: str, password: str,
                     sheet_names: list = None):
    """移除 Sheet 级保护"""
    from openpyxl import load_workbook

    wb = load_workbook(input_file)

    if sheet_names:
        target_sheets = [ws for ws in wb.worksheets if ws.title in sheet_names]
    else:
        target_sheets = wb.worksheets

    for ws in target_sheets:
        if ws.protection.sheet:
            ws.protection.sheet = False
            ws.protection.password = None
            print(f"  🔓 Sheet '{ws.title}' 已解除保护")
        else:
            print(f"  ℹ️ Sheet '{ws.title}' 未被保护")

    wb.save(output_file)
    print(f"\n✅ Sheet 保护已移除: {output_file}")


def main():
    parser = argparse.ArgumentParser(
        description='Excel 密码保护/解密工具'
    )
    parser.add_argument('input_file', help='输入 Excel 文件路径')
    parser.add_argument('--output', help='输出文件路径（不指定则覆盖原文件）')
    parser.add_argument('--mode', required=True,
                        choices=['encrypt', 'decrypt', 'sheet-protect', 'sheet-unprotect'],
                        help='操作模式: encrypt(文件级加密), decrypt(文件级解密), '
                             'sheet-protect(Sheet级保护), sheet-unprotect(移除Sheet保护)')
    parser.add_argument('--password', required=True, help='密码')
    parser.add_argument('--sheet-protect', nargs='*', dest='sheets',
                        help='保护指定 Sheet（不指定则保护所有 Sheet）')
    parser.add_argument('--allow', nargs='*',
                        choices=['select_cells', 'sort', 'filter',
                                 'insert_rows', 'delete_rows', 'format_cells'],
                        help='允许的操作列表')

    args = parser.parse_args()

    # 校验输入
    if not os.path.exists(args.input_file):
        print(f"❌ 输入文件不存在: {args.input_file}")
        sys.exit(1)

    output_file = args.output or args.input_file

    # 确保输出目录存在
    output_dir = os.path.dirname(output_file)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    print(f"📄 输入: {args.input_file}")
    print(f"📄 输出: {output_file}")
    print(f"🔧 模式: {args.mode}\n")

    if args.mode == 'encrypt':
        encrypt_file(args.input_file, output_file, args.password)
    elif args.mode == 'decrypt':
        decrypt_file(args.input_file, output_file, args.password)
    elif args.mode == 'sheet-protect':
        protect_sheets(args.input_file, output_file, args.password,
                       sheet_names=args.sheets, allow_operations=args.allow)
    elif args.mode == 'sheet-unprotect':
        unprotect_sheets(args.input_file, output_file, args.password,
                         sheet_names=args.sheets)


if __name__ == "__main__":
    main()
