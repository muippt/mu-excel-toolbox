#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 工具箱 — 写入/创建 Excel
创建新 Excel 文件或向已有文件追加数据，支持基础样式设置。
"""

import argparse
import sys
import json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from utils import ensure_dependencies, safe_write, ErrorCollector


def load_data(data_source):
    """从多种来源加载数据。

    支持:
    - JSON 字符串
    - JSON 文件路径
    - stdin（当 data_source 为 "-" 或 None 且 stdin 有数据时）

    Args:
        data_source: 数据来源（JSON 字符串、文件路径或 "-"）

    Returns:
        list: 数据行列表（每行是 list 或 dict）
    """
    # 从 stdin 读取
    if data_source == "-" or (data_source is None and not sys.stdin.isatty()):
        raw = sys.stdin.read().strip()
        if not raw:
            raise ValueError("stdin 无数据输入")
        return json.loads(raw)

    if data_source is None:
        raise ValueError("未指定数据来源，请使用 --data 参数或通过 stdin 传入")

    # 尝试作为文件路径
    path = Path(data_source)
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    # 尝试作为 JSON 字符串解析
    try:
        return json.loads(data_source)
    except json.JSONDecodeError:
        raise ValueError(
            f"无法解析数据: '{data_source[:100]}...'\n"
            "请提供有效的 JSON 字符串、JSON 文件路径，或通过 stdin 传入数据"
        )


def create_workbook(data, headers=None, sheet_name="Sheet1",
                    col_width=None, freeze_header=True):
    """创建新的 Excel 工作簿。

    Args:
        data: 数据列表（每项可以是 list 或 dict）
        headers: 列头列表（如为 None 且 data 项为 dict 则自动提取）
        sheet_name: Sheet 名称
        col_width: 列宽（数字或 "auto"）
        freeze_header: 是否冻结首行

    Returns:
        openpyxl.Workbook: 工作簿对象
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 确定表头
    if headers is None and data:
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
        else:
            # 无表头，数据直接写入
            headers = None

    # 写入表头
    start_row = 1
    if headers:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        start_row = 2

    # 写入数据
    for row_idx, row_data in enumerate(data, start_row):
        if isinstance(row_data, dict):
            # dict 数据按 headers 顺序写入
            if headers:
                for col_idx, key in enumerate(headers, 1):
                    value = row_data.get(key, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                for col_idx, (key, value) in enumerate(row_data.items(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        elif isinstance(row_data, (list, tuple)):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            ws.cell(row=row_idx, column=1, value=row_data)

    # 设置列宽
    if col_width == "auto" or col_width is None:
        _auto_col_width(ws)
    elif col_width:
        try:
            width = float(col_width)
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = width
        except (ValueError, TypeError):
            _auto_col_width(ws)

    # 冻结首行
    if freeze_header and headers:
        ws.freeze_panes = "A2"

    return wb


def append_to_workbook(path, data, headers=None, sheet_name=None):
    """向已有 Excel 文件追加数据。

    Args:
        path: 文件路径
        data: 要追加的数据
        headers: 列头（仅在创建新 Sheet 时使用）
        sheet_name: Sheet 名称（默认使用第一个）

    Returns:
        openpyxl.Workbook: 工作簿对象
    """
    from openpyxl import load_workbook

    wb = load_workbook(path)

    if sheet_name:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            # 新 Sheet 写入表头
            if headers:
                from openpyxl.styles import Font, Alignment
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
    else:
        ws = wb.active

    # 追加数据
    for row_data in data:
        if isinstance(row_data, dict):
            # 获取现有表头顺序
            if ws.max_row and ws.max_row >= 1:
                existing_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                row_values = [row_data.get(h, "") for h in existing_headers]
            else:
                row_values = list(row_data.values())
            ws.append(row_values)
        elif isinstance(row_data, (list, tuple)):
            ws.append(list(row_data))
        else:
            ws.append([row_data])

    return wb


def _auto_col_width(ws):
    """自动调整列宽。"""
    from openpyxl.utils import get_column_letter

    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, min(ws.max_row + 1, 101)):  # 只检查前 100 行
            cell = ws.cell(row=row, column=col)
            if cell.value:
                # 计算显示宽度（中文占 2 个字符宽度）
                cell_str = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in cell_str)
                max_length = max(max_length, length)
        # 设置列宽，加 2 留余量，上限 50
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 8)


def main():
    parser = argparse.ArgumentParser(
        description="创建新 Excel 或向已有文件追加数据",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 从 JSON 字符串创建
  python write_excel.py --output result.xlsx --data '[{"姓名":"张三","年龄":30}]'

  # 从 JSON 文件创建
  python write_excel.py --output result.xlsx --data data.json

  # 从 stdin 创建
  echo '[{"name":"test"}]' | python write_excel.py --output result.xlsx --data -

  # 追加数据
  python write_excel.py --output exist.xlsx --data '[{"姓名":"李四"}]' --append

  # 指定表头和 Sheet
  python write_excel.py --output result.xlsx --data '[[1,2],[3,4]]' --headers "A" "B" --sheet "数据"
        """,
    )
    parser.add_argument("--output", "-o", required=True, help="输出文件路径")
    parser.add_argument("--data", "-d", help="JSON 数据、JSON 文件路径或 \"-\" 从 stdin 读取")
    parser.add_argument("--headers", nargs="+", help="列头列表")
    parser.add_argument("--sheet", "-s", default="Sheet1", help="Sheet 名称（默认: Sheet1）")
    parser.add_argument("--append", "-a", action="store_true", help="追加模式（向已有文件追加）")
    parser.add_argument("--col-width", help="列宽（数字或 'auto'，默认自动）")
    parser.add_argument("--no-freeze", action="store_true", help="不冻结首行")

    args = parser.parse_args()

    # 检查依赖
    missing = ensure_dependencies()
    if missing:
        sys.exit(1)

    errors = ErrorCollector()

    try:
        # 加载数据
        data = load_data(args.data)
        if not isinstance(data, list):
            data = [data]

        output_path = Path(args.output)

        if args.append and output_path.exists():
            # 追加模式
            wb = append_to_workbook(
                output_path,
                data,
                headers=args.headers,
                sheet_name=args.sheet,
            )
            result_path = safe_write(wb, output_path)
            print(f"✅ 已追加 {len(data)} 行数据到: {result_path}")
        else:
            # 创建模式
            wb = create_workbook(
                data,
                headers=args.headers,
                sheet_name=args.sheet,
                col_width=args.col_width,
                freeze_header=not args.no_freeze,
            )
            result_path = safe_write(wb, output_path)
            print(f"✅ 已创建 Excel 文件: {result_path}")
            print(f"   数据行数: {len(data)}，Sheet: {args.sheet}")

    except Exception as e:
        print(f"❌ 写入失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
