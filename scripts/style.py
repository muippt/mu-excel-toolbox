#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 样式排版 + 条件格式工具
功能：为 Excel 添加样式美化和条件格式
- 样式模式：表头加粗、背景色、边框、自适应列宽、冻结行、字体设置
- 条件格式：基于规则的单元格高亮、色阶、数据条
"""

import argparse
import os
import sys
import json
from copy import copy
from typing import Optional, List

# 添加脚本所在目录到 path，以便引用 utils
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import detect_file_type, safe_read, format_output, ErrorCollector

try:
    from openpyxl import load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment,
        numbers
    )
    from openpyxl.formatting.rule import (
        CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
    )
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("❌ 缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)


def col_letter_to_index(col: str) -> int:
    """列字母转为 1-based 索引，支持字母或数字输入"""
    try:
        return int(col)
    except ValueError:
        return column_index_from_string(col.upper())


def apply_style(wb, args, collector: ErrorCollector):
    """应用样式模式"""
    ws = wb.active

    # 表头加粗
    if args.header_bold:
        for cell in ws[1]:
            cell.font = Font(
                bold=True,
                name=args.font if args.font else cell.font.name,
                size=args.font_size if args.font_size else cell.font.size
            )

    # 表头背景色
    if args.header_bg:
        color = args.header_bg.lstrip('#')
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for cell in ws[1]:
            cell.font = Font(
                bold=cell.font.bold,
                name=args.font if args.font else cell.font.name,
                size=args.font_size if args.font_size else cell.font.size,
                color='FFFFFF' if _is_dark_color(color) else cell.font.color
            )
            cell.fill = fill

    # 字体设置（全局，非表头行）
    if args.font or args.font_size:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(
                    name=args.font if args.font else cell.font.name,
                    size=args.font_size if args.font_size else cell.font.size,
                    bold=cell.font.bold
                )

    # 边框
    if args.border and args.border != 'none':
        thin_side = Side(style='thin')
        if args.border == 'all':
            border = Border(
                left=thin_side, right=thin_side,
                top=thin_side, bottom=thin_side
            )
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
        elif args.border == 'outer':
            # 只加外边框
            max_row = ws.max_row
            max_col = ws.max_column
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    sides = {}
                    if row_idx == 1:
                        sides['top'] = thin_side
                    if row_idx == max_row:
                        sides['bottom'] = thin_side
                    if col_idx == 1:
                        sides['left'] = thin_side
                    if col_idx == max_col:
                        sides['right'] = thin_side
                    if sides:
                        ws.cell(row=row_idx, column=col_idx).border = Border(**sides)

    # 自适应列宽
    if args.auto_width:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[:101]:  # 只检查前 100 行，避免大文件性能问题
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    # 中文字符算 2 个宽度
                    length = sum(2 if ord(c) > 127 else 1 for c in val)
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    # 冻结行
    if args.freeze_row:
        ws.freeze_panes = f'A{args.freeze_row + 1}'

    return wb


def apply_conditional(wb, args, collector: ErrorCollector):
    """应用条件格式"""
    ws = wb.active
    max_row = ws.max_row

    # 解析条件格式规则（支持多列多规则，通过 JSON 或单个参数）
    rules = []

    if args.rules_json:
        # JSON 格式批量规则
        try:
            rules = json.loads(args.rules_json)
        except json.JSONDecodeError as e:
            collector.add_error(f"JSON 规则解析失败: {e}")
            return wb
    elif args.column and args.rule:
        # 单条规则
        rules.append({
            'column': args.column,
            'rule': args.rule,
            'value': args.value,
            'color': args.color or 'FF0000',
            'font_color': args.font_color,
        })

    # 色阶
    if args.color_scale:
        cols = args.color_scale.split(',')
        for col in cols:
            col_idx = col_letter_to_index(col.strip())
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{max_row}"
            rule = ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
            ws.conditional_formatting.add(cell_range, rule)

    # 数据条
    if args.data_bar:
        cols = args.data_bar.split(',')
        for col in cols:
            col_idx = col_letter_to_index(col.strip())
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{max_row}"
            rule = DataBarRule(
                start_type='min', end_type='max',
                color='638EC6', showValue=True
            )
            ws.conditional_formatting.add(cell_range, rule)

    # 处理规则列表
    for r in rules:
        col = r.get('column')
        rule_type = r.get('rule')
        value = r.get('value')
        fill_color = r.get('color', 'FF0000').lstrip('#')
        font_color = r.get('font_color')

        if not col or not rule_type:
            collector.add_warning("规则缺少 column 或 rule 字段，已跳过")
            continue

        col_idx = col_letter_to_index(col)
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{max_row}"

        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        font_kwargs = {}
        if font_color:
            font_kwargs['color'] = font_color.lstrip('#')
        font = Font(**font_kwargs) if font_kwargs else None

        # 构造 openpyxl 条件格式规则
        rule_kwargs = {'fill': fill}
        if font:
            rule_kwargs['font'] = font

        operator_map = {
            'gt': 'greaterThan',
            'lt': 'lessThan',
            'eq': 'equal',
            'gte': 'greaterThanOrEqual',
            'lte': 'lessThanOrEqual',
        }

        if rule_type in operator_map:
            cf_rule = CellIsRule(
                operator=operator_map[rule_type],
                formula=[str(value)],
                **rule_kwargs
            )
            ws.conditional_formatting.add(cell_range, cf_rule)

        elif rule_type == 'between':
            # value 格式: "60,90"
            parts = str(value).split(',')
            if len(parts) != 2:
                collector.add_error(f"between 规则需要两个值(逗号分隔)，实际: {value}")
                continue
            cf_rule = CellIsRule(
                operator='between',
                formula=[parts[0].strip(), parts[1].strip()],
                **rule_kwargs
            )
            ws.conditional_formatting.add(cell_range, cf_rule)

        elif rule_type == 'contains':
            # 使用公式实现包含
            formula = f'NOT(ISERROR(SEARCH("{value}",{col_letter}2)))'
            cf_rule = FormulaRule(
                formula=[formula],
                **rule_kwargs
            )
            ws.conditional_formatting.add(cell_range, cf_rule)

        elif rule_type == 'top_n':
            # 前 N 名高亮
            n = int(value) if value else 10
            formula = f'{col_letter}2>=LARGE(${col_letter}$2:${col_letter}${max_row},{n})'
            cf_rule = FormulaRule(
                formula=[formula],
                **rule_kwargs
            )
            ws.conditional_formatting.add(cell_range, cf_rule)

        elif rule_type == 'bottom_n':
            # 后 N 名高亮
            n = int(value) if value else 10
            formula = f'{col_letter}2<=SMALL(${col_letter}$2:${col_letter}${max_row},{n})'
            cf_rule = FormulaRule(
                formula=[formula],
                **rule_kwargs
            )
            ws.conditional_formatting.add(cell_range, cf_rule)

        else:
            collector.add_warning(f"不支持的规则类型: {rule_type}")

    return wb


def _is_dark_color(hex_color: str) -> bool:
    """判断颜色是否为深色（用于自动切换字体颜色）"""
    try:
        hex_color = hex_color.lstrip('#')
        if len(hex_color) == 8:
            hex_color = hex_color[2:]  # 去掉 alpha
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        # 使用亮度公式
        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        return luminance < 0.5
    except Exception:
        return False


def main():
    parser = argparse.ArgumentParser(
        description='Excel 样式排版 + 条件格式工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 样式模式：表头加粗加背景色，自适应列宽，加边框
  python style.py data.xlsx --mode style --header-bold --header-bg 4472C4 --auto-width --border all

  # 条件格式：绩效低于 60 标红
  python style.py data.xlsx --mode conditional --column D --rule lt --value 60 --color FF0000

  # 条件格式：色阶
  python style.py data.xlsx --mode conditional --color-scale D,E

  # 批量条件格式（JSON）
  python style.py data.xlsx --mode conditional --rules-json '[{"column":"D","rule":"lt","value":"60","color":"FF0000"},{"column":"D","rule":"gt","value":"90","color":"00B050"}]'
"""
    )

    parser.add_argument('input_file', help='输入 Excel 文件路径')
    parser.add_argument('--output', '-o', help='输出文件路径（默认覆盖原文件）')
    parser.add_argument('--mode', choices=['style', 'conditional'], default='style',
                        help='工作模式: style=样式排版, conditional=条件格式')

    # 样式模式参数
    style_group = parser.add_argument_group('样式模式参数')
    style_group.add_argument('--header-bold', action='store_true', help='表头加粗')
    style_group.add_argument('--header-bg', help='表头背景色（十六进制，如 4472C4）')
    style_group.add_argument('--border', choices=['all', 'outer', 'none'], help='边框样式')
    style_group.add_argument('--auto-width', action='store_true', help='自适应列宽')
    style_group.add_argument('--freeze-row', type=int, help='冻结到第 N 行（冻结该行及以上）')
    style_group.add_argument('--font', help='字体名称（如 微软雅黑）')
    style_group.add_argument('--font-size', type=int, help='字号')

    # 条件格式参数
    cond_group = parser.add_argument_group('条件格式参数')
    cond_group.add_argument('--column', help='目标列（字母或数字）')
    cond_group.add_argument('--rule',
                           choices=['gt', 'lt', 'eq', 'gte', 'lte', 'between', 'contains', 'top_n', 'bottom_n'],
                           help='条件规则')
    cond_group.add_argument('--value', help='条件值（between 用逗号分隔两个值）')
    cond_group.add_argument('--color', help='条件满足时的填充色（十六进制）')
    cond_group.add_argument('--font-color', help='条件满足时的字体色（十六进制）')
    cond_group.add_argument('--rules-json', help='批量条件规则（JSON 数组格式）')
    cond_group.add_argument('--color-scale', help='渐变色阶应用的列（逗号分隔，如 D,E）')
    cond_group.add_argument('--data-bar', help='数据条应用的列（逗号分隔，如 D,E）')

    args = parser.parse_args()

    # 验证输入文件
    if not os.path.exists(args.input_file):
        print(f"❌ 文件不存在: {args.input_file}")
        sys.exit(1)

    collector = ErrorCollector()

    # 加载工作簿
    try:
        wb = load_workbook(args.input_file)
    except Exception as e:
        print(f"❌ 无法打开文件: {e}")
        sys.exit(1)

    # 根据模式执行
    if args.mode == 'style':
        wb = apply_style(wb, args, collector)
    elif args.mode == 'conditional':
        wb = apply_conditional(wb, args, collector)

    # 保存
    output_path = args.output or args.input_file
    try:
        wb.save(output_path)
        print(f"✅ 已保存到: {output_path}")
    except Exception as e:
        print(f"❌ 保存失败: {e}")
        sys.exit(1)

    # 输出错误报告
    if collector.has_errors() or collector.has_warnings():
        print(collector.report())


if __name__ == "__main__":
    main()
